from openpyxl import Workbook, load_workbook
from pandas_datareader import data as web
from tqdm import trange
import datetime

today = datetime.date.today()
today = f'{today.month}-{today.day}-{today.year}'

# wbPath is the path to your excel file 
# make sure it's right
wbPath = 'path to the Workbook'
wb = load_workbook(wbPath)
ws = wb.active

# variables about where the stocks tickers are and where the price will be
startRow = 1   # the first row with a stock ticker
endRow = 10    # the last row with a stock ticker
stocksColumn = 'A' # the column where the tickers are
priceColumn = 'B'  # the column where the prices will be

for n in trange(startRow, endRow):
	# look on the worksheet the stock ticker
	stock = ws[stocksColumn + str(n)].value
	
	# gets the stock price from yahoo finance
	price = web.DataReader(stock, data_source='yahoo', start=today, end=today)
	price = round(price['Close'][today], 2)  # rounding the number

	# put the price on the worksheet
	ws[priceColumn + str(n)].value = price

# saves the file
wb.save(wbPath)